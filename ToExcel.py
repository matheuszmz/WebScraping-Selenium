import openpyxl
import os

def check_values_in_list(list_values):
    i = []
    for v in list_values:
        if (len(v) not in i):
            i.append(len(v))
    return len(i)

def export_to_excel(path=os.getcwd(), dictionary):
    if (type(dictionary) is not dict):
        print('Erro - Dicionário inválido.')
    else:
        try:
            keys = list(dictionary.keys())
            values = list(dictionary.values())
            if (len(keys) == len(values) and check_values_in_list(values) == 1):
                i, j = 2, 1
                wb = openpyxl.Workbook()
                ws = wb.active
                for key in keys:
                    ws.cell(row=1, column=j).value = key
                    j +=1
                j = 1
                for value in values:
                    for v in value:
                        ws.cell(row=i, column=j).value = v
                        i += 1
                    j += 1
                    i = 2
                wb.save(path)        
            else:
                print('Erro - As chaves não possuem a mesma quantidade de valores.')
        except PermissionError:
            print('Acesso negado ao {}'.format(path))
